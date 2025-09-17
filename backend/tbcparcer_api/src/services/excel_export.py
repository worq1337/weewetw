import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import io
from typing import List, Dict, Optional, Any

class ExcelExportService:
    def __init__(self):
        # Конфигурация колонок соответствует таблице на сайте (ТЗ)
        self.column_config = [
            {
                'key': 'receipt_number',
                'header': 'Номер чека',
                'width': 18,
                'alignment': 'center'
            },
            {
                'key': 'date_time',
                'header': 'Дата и время',
                'width': 22,
                'alignment': 'center',
                'number_format': 'dd.mm.yyyy hh:mm'
            },
            {
                'key': 'day_name',
                'header': 'Д.н.',
                'width': 9,
                'alignment': 'center'
            },
            {
                'key': 'date',
                'header': 'Дата',
                'width': 14,
                'alignment': 'center',
                'number_format': 'dd.mm.yyyy'
            },
            {
                'key': 'time',
                'header': 'Время',
                'width': 12,
                'alignment': 'center',
                'number_format': 'hh:mm'
            },
            {
                'key': 'operator_seller',
                'header': 'Оператор/Продавец',
                'width': 22
            },
            {
                'key': 'application',
                'header': 'Приложение',
                'width': 20
            },
            {
                'key': 'amount',
                'header': 'Сумма',
                'width': 18,
                'alignment': 'right',
                'number_format': '#,##0.00'
            },
            {
                'key': 'balance',
                'header': 'Остаток',
                'width': 18,
                'alignment': 'right',
                'number_format': '#,##0.00'
            },
            {
                'key': 'card_number',
                'header': 'ПК',
                'width': 12,
                'alignment': 'center'
            },
            {
                'key': 'p2p',
                'header': 'P2P',
                'width': 10,
                'alignment': 'center'
            },
            {
                'key': 'transaction_type',
                'header': 'Тип транзакции',
                'width': 20
            },
            {
                'key': 'currency',
                'header': 'Валюта',
                'width': 12,
                'alignment': 'center'
            },
            {
                'key': 'data_source',
                'header': 'Источник данных',
                'width': 18
            },
            {
                'key': 'category',
                'header': 'Категория',
                'width': 18
            }
        ]
        
        self.operation_types = {
            'payment': 'Оплата',
            'refill': 'Пополнение',
            'conversion': 'Конверсия',
            'cancel': 'Отмена'
        }
    
    def _get_day_of_week(self, date_time):
        """Получить день недели на русском"""
        if not date_time:
            return ""
        
        if isinstance(date_time, str):
            try:
                date_time = datetime.fromisoformat(date_time.replace('Z', '+00:00'))
            except:
                return ""
        
        days = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
        return days[date_time.weekday()]
    
    def _format_card_number(self, card_number):
        """Форматировать номер карты для колонки ПК"""
        if not card_number:
            return ""
        # Если уже есть *, возвращаем как есть, иначе добавляем *
        if card_number.startswith('*'):
            return card_number
        return f"*{card_number[-4:]}" if len(card_number) >= 4 else card_number
    
    def _get_transaction_data(self, transaction: Dict[str, Any]) -> Dict[str, Any]:
        """Преобразовать транзакцию в данные для Excel согласно ТЗ"""
        date_time_obj = None
        if transaction.get('date_time'):
            if isinstance(transaction['date_time'], str):
                try:
                    date_time_obj = datetime.fromisoformat(transaction['date_time'].replace('Z', '+00:00'))
                    if date_time_obj.tzinfo is not None:
                        date_time_obj = date_time_obj.replace(tzinfo=None)
                except Exception:
                    date_time_obj = None
            else:
                date_time_obj = transaction['date_time']
                if isinstance(date_time_obj, datetime) and date_time_obj.tzinfo is not None:
                    date_time_obj = date_time_obj.replace(tzinfo=None)

        # Генерируем номер чека (если нет в данных)
        receipt_number = transaction.get('receipt_number')
        if not receipt_number:
            receipt_number = f"CHK{str(transaction.get('id', '000')).zfill(3)}"

        # Определяем P2P (пока заглушка, так как в модели нет этого поля)
        p2p_value = 'Нет'  # По умолчанию
        if transaction.get('operation_type') == 'conversion':
            p2p_value = 'Да'

        date_value: Optional[date] = None
        time_value: Optional[time] = None
        if isinstance(date_time_obj, datetime):
            date_value = date_time_obj.date()
            time_value = date_time_obj.time().replace(second=0, microsecond=0)

        return {
            'receipt_number': receipt_number,
            'date_time': date_time_obj if date_time_obj else None,
            'day_name': self._get_day_of_week(date_time_obj),
            'date': date_value,
            'time': time_value,
            'operator_seller': transaction.get('operator_name', ''),
            'application': transaction.get('operator_description', ''),
            'amount': transaction.get('amount', 0),
            'balance': transaction.get('balance', 0),
            'card_number': self._format_card_number(transaction.get('card_number', '')),
            'p2p': p2p_value,
            'transaction_type': self.operation_types.get(
                transaction.get('operation_type', ''),
                transaction.get('operation_type', '')
            ),
            'currency': transaction.get('currency', 'UZS'),
            'data_source': transaction.get('data_source', 'API'),
            'category': self._get_category_from_description(transaction.get('description', ''))
        }
    
    def _get_category_from_description(self, description):
        """Определить категорию по описанию"""
        if not description:
            return 'Прочее'
        
        description_lower = description.lower()
        if 'перевод' in description_lower or 'конверсия' in description_lower:
            return 'Переводы'
        elif 'оплата' in description_lower:
            return 'Покупки'
        elif 'пополнение' in description_lower:
            return 'Пополнения'
        else:
            return 'Прочее'
    
    def export_transactions(self, transactions: List[Dict[str, Any]]) -> io.BytesIO:
        """
        Экспорт транзакций в Excel с точным соответствием таблице сайта
        
        Args:
            transactions: Список транзакций
        
        Returns:
            BytesIO объект с Excel файлом
        """
        # Создаем новую книгу
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Финансовые транзакции"
        
        # Стиль границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alignments = {
            'left': Alignment(horizontal='left', vertical='center'),
            'center': Alignment(horizontal='center', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }

        # Создаем заголовки (строка 1)
        for col_idx, column in enumerate(self.column_config, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column['header'])
            # Стиль заголовков
            cell.font = Font(bold=True, size=12)
            cell.alignment = alignments['center']
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            cell.border = thin_border

        # Заполняем данные транзакций
        for row_idx, transaction in enumerate(transactions, 2):  # Начинаем с 2-й строки
            transaction_data = self._get_transaction_data(transaction)

            for col_idx, column in enumerate(self.column_config, 1):
                value = transaction_data.get(column['key'])
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                alignment_key = column.get('alignment', 'left')
                cell.alignment = alignments.get(alignment_key, alignments['left'])

                number_format = column.get('number_format')
                if number_format:
                    cell.number_format = number_format
                else:
                    if isinstance(value, datetime):
                        cell.number_format = 'dd.mm.yyyy hh:mm'
                    elif isinstance(value, date):
                        cell.number_format = 'dd.mm.yyyy'
                    elif isinstance(value, time):
                        cell.number_format = 'hh:mm'

                # Границы для всех ячеек
                cell.border = thin_border

        # Устанавливаем ширину колонок
        for col_idx, column in enumerate(self.column_config, 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = column['width']
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output

