import io
from datetime import time

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.services.ai_parser import AIParsingService
from src.services.excel_export import ExcelExportService


@pytest.fixture(scope='module')
def parsing_service():
    """Provide a single parsing service instance for the module."""
    return AIParsingService()


def _build_receipt(date_str: str, amount: str, balance: str, card: str, operator: str, description: str) -> str:
    return (
        f"Дата: {date_str}\n"
        f"Сумма: {amount} UZS\n"
        f"Баланс: {balance} UZS\n"
        f"Карта: {card}\n"
        f"Оператор: {operator}\n"
        f"Описание: {description}\n"
    )


def test_parser_to_export_flow(client, parsing_service):
    telegram_id = 777000
    username = 'e2e-tester'

    # Создаем пользователя (этап UI регистрации/авторизации)
    user_response = client.post('/api/users', json={'telegram_id': telegram_id, 'username': username})
    assert user_response.status_code in (200, 201)
    created_user = user_response.get_json()
    assert created_user['telegram_id'] == telegram_id

    receipts = [
        {
            'text': _build_receipt(
                '2024-05-12 08:45',
                '125000',
                '980000',
                '*1234',
                'Humans',
                'Оплата мобильной связи'
            ),
            'expected_operation': 'payment',
            'expected_amount': 125000.0,
        },
        {
            'text': _build_receipt(
                '2024-05-10 17:20',
                '350000',
                '1240000',
                '*9876',
                'Milliy 2.0',
                'Пополнение карты'
            ),
            'expected_operation': 'refill',
            'expected_amount': 350000.0,
        },
    ]

    saved_transaction_ids = []

    # Парсим и сохраняем транзакции через API (parser -> API/DB)
    for receipt in receipts:
        preview = parsing_service.parse_receipt(receipt['text'])
        assert preview.get('operation_type') == receipt['expected_operation']
        assert pytest.approx(float(preview.get('amount', 0))) == receipt['expected_amount']

        response = client.post('/api/ai/parse-and-save', json={'text': receipt['text'], 'telegram_id': telegram_id})
        assert response.status_code == 200
        payload = response.get_json()
        transaction = payload['transaction']
        saved_transaction_ids.append(transaction['id'])

        assert transaction['operation_type'] == receipt['expected_operation']
        assert pytest.approx(transaction['amount']) == receipt['expected_amount']
        assert transaction['data_source'] == 'AI Parser'

    assert len(set(saved_transaction_ids)) == len(receipts)

    # Проверяем навигацию (пагинация) и фильтр по пользователю (UI слой)
    list_page_1 = client.get('/api/transactions', query_string={'telegram_id': telegram_id, 'per_page': 1, 'page': 1})
    assert list_page_1.status_code == 200
    payload_page_1 = list_page_1.get_json()
    assert payload_page_1['total'] == len(receipts)
    assert payload_page_1['pages'] == len(receipts)
    assert payload_page_1['current_page'] == 1
    assert len(payload_page_1['transactions']) == 1
    assert payload_page_1['transactions'][0]['id'] in saved_transaction_ids

    list_page_2 = client.get('/api/transactions', query_string={'telegram_id': telegram_id, 'per_page': 1, 'page': 2})
    assert list_page_2.status_code == 200
    payload_page_2 = list_page_2.get_json()
    assert payload_page_2['current_page'] == 2
    assert len(payload_page_2['transactions']) == 1
    assert payload_page_2['transactions'][0]['id'] in saved_transaction_ids
    page_transaction_ids = {payload_page_1['transactions'][0]['id'], payload_page_2['transactions'][0]['id']}
    assert page_transaction_ids == set(saved_transaction_ids)

    # Проверяем поведение при отсутствии telegram_id (валидация фильтра)
    missing_filter_response = client.get('/api/transactions')
    assert missing_filter_response.status_code == 400

    # Экспортируем данные и проверяем формат Excel (UI -> export)
    export_response = client.post('/api/export/excel', json={'telegram_id': telegram_id})
    assert export_response.status_code == 200

    workbook = load_workbook(filename=io.BytesIO(export_response.data))
    worksheet = workbook.active

    export_service = ExcelExportService()
    expected_headers = [column['header'] for column in export_service.column_config]
    actual_headers = [cell.value for cell in worksheet[1]]
    assert actual_headers == expected_headers

    # Проверяем ширину колонок и числовой формат времени
    for idx, column in enumerate(export_service.column_config, start=1):
        column_letter = get_column_letter(idx)
        assert worksheet.column_dimensions[column_letter].width == pytest.approx(column['width'])

    time_cell = worksheet.cell(row=2, column=5)
    assert isinstance(time_cell.value, time)
    assert time_cell.number_format == 'hh:mm'

    # Дополнительно проверяем, что все строки выгрузки соответствуют числу транзакций
    assert worksheet.max_row == len(receipts) + 1
